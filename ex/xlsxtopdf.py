# xlsxtopdf.py
# program to conver xlsx file to pdf

from openpyxl import load_workbook
from PDFWriter import PDFWriter

workbook = load_workbook('test.xlsx', guess_types=True, data_only=True)
worksheet = workbook.active

pw = PDFWriter('test.pdf')
pw.setFont('Courier', 12)
pw.setHeader('xlsxtopdf.py - convert xlsx file to pdf')
pw.setFooter('generated using openpyxl and xtopdf')

ws_range = worksheet.iter_rows('A1:H13')
for row in ws_range:
    s = ''
    for cell in row:
        if cell.value is None:
            s += ' ' * 11
        else:
            s += str(cell.value).rjust(10) + ' '
        pw.writeLine(s)

pw.savePage()
pw.close()
